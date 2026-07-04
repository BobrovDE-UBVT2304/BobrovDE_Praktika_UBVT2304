# reports/report_generator.py
import json
import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns
from fpdf import FPDF
import os
from datetime import datetime

class ReportGenerator:
    def __init__(self, results_path='reports/results.json'):
        with open(results_path, 'r') as f:
            self.results = json.load(f)
        self.output_dir = 'reports'
        os.makedirs(self.output_dir, exist_ok=True)
    
    def generate_pdf_report(self):
        """Генерация PDF отчета"""
        pdf = FPDF()
        pdf.add_page()
        
        # Заголовок
        pdf.set_font('Arial', 'B', 20)
        pdf.cell(200, 20, 'Traffic Sign Recognition Report', ln=True, align='C')
        pdf.set_font('Arial', '', 12)
        pdf.cell(200, 10, f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M")}', ln=True, align='C')
        pdf.ln(10)
        
        # Информация о моделях
        pdf.set_font('Arial', 'B', 14)
        pdf.cell(200, 10, 'Model Comparison', ln=True)
        pdf.ln(5)
        
        # Таблица сравнения
        df = pd.DataFrame(self.results).T.reset_index().rename(columns={'index': 'Model'})
        
        # Таблица в PDF
        pdf.set_font('Arial', 'B', 10)
        pdf.cell(40, 10, 'Model', 1)
        pdf.cell(25, 10, 'Accuracy', 1)
        pdf.cell(25, 10, 'Precision', 1)
        pdf.cell(25, 10, 'Recall', 1)
        pdf.cell(25, 10, 'F1 Score', 1)
        pdf.cell(25, 10, 'Size (M)', 1)
        pdf.cell(30, 10, 'Time (ms)', 1)
        pdf.ln()
        
        pdf.set_font('Arial', '', 9)
        for _, row in df.iterrows():
            pdf.cell(40, 10, str(row['Model']), 1)
            pdf.cell(25, 10, f"{row.get('accuracy', 0):.3f}", 1)
            pdf.cell(25, 10, f"{row.get('precision', 0):.3f}", 1)
            pdf.cell(25, 10, f"{row.get('recall', 0):.3f}", 1)
            pdf.cell(25, 10, f"{row.get('f1_score', 0):.3f}", 1)
            pdf.cell(25, 10, f"{row.get('model_size_mb', 0):.1f}", 1)
            pdf.cell(30, 10, f"{row.get('avg_inference_time_ms', 0):.1f}", 1)
            pdf.ln()
        
        # Лучшая модель
        pdf.ln(10)
        best_model = df.loc[df['accuracy'].idxmax()]
        pdf.set_font('Arial', 'B', 12)
        pdf.cell(200, 10, f"Best Model: {best_model['Model']}", ln=True)
        pdf.set_font('Arial', '', 11)
        pdf.cell(200, 10, f"Accuracy: {best_model['accuracy']:.3f}", ln=True)
        pdf.cell(200, 10, f"F1 Score: {best_model['f1_score']:.3f}", ln=True)
        pdf.cell(200, 10, f"Inference Time: {best_model['avg_inference_time_ms']:.1f}ms", ln=True)
        
        # Сохранение PDF
        pdf_path = os.path.join(self.output_dir, 'full_report.pdf')
        pdf.output(pdf_path)
        print(f"PDF report saved to {pdf_path}")
        return pdf_path
    
    def generate_excel_report(self):
        """Генерация Excel отчета"""
        df = pd.DataFrame(self.results).T.reset_index().rename(columns={'index': 'Model'})
        
        excel_path = os.path.join(self.output_dir, 'full_report.xlsx')
        with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Model Comparison', index=False)
            
            # Создание графика и добавление в Excel
            fig, axes = plt.subplots(2, 2, figsize=(12, 10))
            
            metrics = ['accuracy', 'precision', 'recall', 'f1_score']
            for idx, metric in enumerate(metrics):
                ax = axes[idx // 2, idx % 2]
                df_sorted = df.sort_values(metric, ascending=False)
                ax.bar(df_sorted['Model'], df_sorted[metric])
                ax.set_title(metric.capitalize())
                ax.set_xticklabels(df_sorted['Model'], rotation=45, ha='right')
                ax.set_ylabel(metric.capitalize())
            
            plt.tight_layout()
            plt.savefig('temp_chart.png', dpi=150, bbox_inches='tight')
            plt.close()
            
            # Вставка изображения в Excel
            from openpyxl.drawing.image import Image
            from openpyxl.utils.dataframe import dataframe_to_rows
            from openpyxl import Workbook
            
            wb = writer.book
            ws = wb['Model Comparison']
            
            # Добавление изображения
            img = Image('temp_chart.png')
            img.width = 800
            img.height = 600
            ws.add_image(img, 'H2')
            
            # Настройка ширины колонок
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 30)
                ws.column_dimensions[column].width = adjusted_width
        
        os.remove('temp_chart.png')
        print(f"Excel report saved to {excel_path}")
        return excel_path
    
    def generate_summary(self):
        """Генерация краткого summary"""
        df = pd.DataFrame(self.results).T.reset_index().rename(columns={'index': 'Model'})
        
        summary = {
            'total_models': len(df),
            'best_model': df.loc[df['accuracy'].idxmax()]['Model'],
            'best_accuracy': df['accuracy'].max(),
            'best_f1': df['f1_score'].max(),
            'fastest_model': df.loc[df['avg_inference_time_ms'].idxmin()]['Model'],
            'fastest_time': df['avg_inference_time_ms'].min(),
            'smallest_model': df.loc[df['model_size_mb'].idxmin()]['Model'],
            'smallest_size': df['model_size_mb'].min(),
            'average_accuracy': df['accuracy'].mean(),
            'average_f1': df['f1_score'].mean()
        }
        
        summary_path = os.path.join(self.output_dir, 'summary.json')
        with open(summary_path, 'w') as f:
            json.dump(summary, f, indent=2)
        
        print("Summary generated:")
        for key, value in summary.items():
            print(f"  {key}: {value}")
        
        return summary

if __name__ == "__main__":
    # Генерация отчетов
    generator = ReportGenerator()
    generator.generate_pdf_report()
    generator.generate_excel_report()
    generator.generate_summary()